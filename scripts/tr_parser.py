from __future__ import annotations
import psycopg
# from psycopg2.extras import execute_batch
from typing import Dict, Any, List, Union, Optional
import requests
from psycopg import connect
from openpyxl import load_workbook
from dotenv import load_dotenv
import os
import time


HEADERS = {
    'Accept': '*/*',
    'Accept-Encoding': 'gzip, deflate, br, zstd',
    'Accept-Language': 'ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7',
    'Cache-Control': 'no-cache',
    'Connection': 'keep-alive',
    'Cookie': 'csrftoken=hzKhFIr0zo7suct4AFwDZxo8IbADeVtw; _ym_uid=1755597274158591030; _ym_d=1755597274; _upassist_client_id=7d20e91a-f305-4462-ace9-d5340e48143a; exitModal=closed; _ym_isad=1; _upassist_session_id=79877a27-e341-462e-8bf5-a46d25cb8ff2; _ym_visorc=w; _cmg_csstM9Z_d=1758973205; _comagic_idM9Z_d=9851790517.13909786757.1758973204',
    'Host': 'truboproduct.ru',
    'Pragma': 'no-cache',
    'Referer': 'https://truboproduct.ru/catalog/bolty/',
    'Sec-Fetch-Dest': 'empty',
    'Sec-Fetch-Mode': 'cors',
    'Sec-Fetch-Site': 'same-origin',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/140.0.0.0 Safari/537.36',
    'sec-ch-ua': '"Chromium";v="140", "Not=A?Brand";v="24", "Google Chrome";v="140"',
    'sec-ch-ua-mobile': '?0',
    'sec-ch-ua-platform': '"Windows"'
}

def open_db():
    load_dotenv()
    database_url = os.getenv("DATABASE_URL")
    if not database_url:
        raise ValueError("DATABASE_URL not found in environment variables")
    conn = psycopg.connect(database_url)
    conn.autocommit = True
    return conn

def get_data():
    url = "https://truboproduct.ru/api/catalog/"
    response = requests.get(url,headers=HEADERS)
    return response.json()

def flatten_tree(
    data: Union[Dict[str, Any], List[Dict[str, Any]]]
) -> List[Dict[str, Any]]:
    """Рекурсивно разворачивает JSON-дерево в список словарей."""
    roots = data if isinstance(data, list) else [data]
    rows: List[Dict[str, Any]] = []

    def walk(node: Dict[str, Any], path_names: List[str], path_slugs: List[str], level: int):
        children = node.get("children") or []
        child_count = len(children)
        is_leaf = 1 if child_count == 0 else 0

        cur_names = path_names + [str(node.get("name", ""))]
        cur_slugs = path_slugs + [str(node.get("slug", ""))]

        row = {
            "id": node.get("id"),
            "parent_id": node.get("parent_id"),
            "name": node.get("name"),
            "slug": node.get("slug"),
            "root_": node.get("root_"),
            "level": level,
            "path_name": "/".join(cur_names),
            "path_slug": "/".join(cur_slugs),
            "is_leaf": bool(is_leaf),
            "child_count": child_count,
        }
        rows.append(row)

        for ch in children:
            walk(ch, cur_names, cur_slugs, level + 1)

    for r in roots:
        walk(r, [], [], 0)

    return rows


def upsert_categories_from_json(
    data: Union[Dict[str, Any], List[Dict[str, Any]]]
):
    """Разворачивает JSON и грузит в PostgreSQL с апсертом по id."""
    rows = flatten_tree(data)

    cols = [
        "id", "parent_id", "name", "slug", "root_",
        "level", "path_name", "path_slug", "is_leaf", "child_count"
    ]

    placeholders = ", ".join(["%s"] * len(cols))
    colnames = ", ".join(cols)
    update_set = ", ".join([f"{c}=EXCLUDED.{c}" for c in cols if c != "id"])

    sql = f"""
    INSERT INTO tr_categories ({colnames})
    VALUES ({placeholders})
    ON CONFLICT (id) DO UPDATE
      SET {update_set};
    """

    values = [[r[c] for c in cols] for r in rows]
    with open_db() as conn:
        with conn.cursor() as cur:
            cur.executemany(sql, values)
        conn.commit()

def get_xlsx(slug, max_retries=3):
    url = f"https://truboproduct.ru/price-list/{slug}/"
    headers = {
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
        'Accept-Encoding': 'gzip, deflate, br, zstd',
        'Accept-Language': 'ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7',
        'Cache-Control': 'no-cache',
        'Connection': 'keep-alive',
        'Cookie': '_upassist_session_id=4c7e43c4-fd7f-48d6-b8fb-ccbb7a490cdc; exitModal=closed; csrftoken=hzKhFIr0zo7suct4AFwDZxo8IbADeVtw; _ym_uid=1755597274158591030; _ym_d=1755597274; _upassist_client_id=7d20e91a-f305-4462-ace9-d5340e48143a; _ym_isad=1; _ym_visorc=w; _cmg_csstM9Z_d=1758973205; _comagic_idM9Z_d=9851790517.13909786757.1758973204',
        'Host': 'truboproduct.ru',
        'Pragma': 'no-cache',
        'Referer': url,
        'Sec-Fetch-Dest': 'document',
        'Sec-Fetch-Mode': 'navigate',
        'Sec-Fetch-Site': 'same-origin',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/140.0.0.0 Safari/537.36',
        'sec-ch-ua': '"Chromium";v="140", "Not=A?Brand";v="24", "Google Chrome";v="140"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"',
        'sec-fetch-user': '?1',
        'upgrade-insecure-requests': '1'
    }

    filename = f"{slug}.xlsx"

    for attempt in range(max_retries):
        try:
            response = requests.get(url, allow_redirects=True, headers=headers, timeout=30)
            response.raise_for_status()  # выбросит исключение, если не 200

            with open(filename, "wb") as f:
                f.write(response.content)

            print(f"Файл сохранён как {filename}")
            return True
            
        except requests.exceptions.SSLError as e:
            print(f"SSL ошибка при попытке {attempt + 1}/{max_retries}: {e}")
            if attempt < max_retries - 1:
                wait_time = (2 ** attempt) * 10  # 10, 20, 40 секунд
                print(f"Ждем {wait_time} секунд перед повтором...")
                time.sleep(wait_time)
            else:
                raise e
        except Exception as e:
            print(f"Ошибка при попытке {attempt + 1}/{max_retries}: {e}")
            if attempt < max_retries - 1:
                wait_time = (2 ** attempt) * 5  # 5, 10, 20 секунд
                print(f"Ждем {wait_time} секунд перед повтором...")
                time.sleep(wait_time)
            else:
                raise e

def upsert_products_from_xlsx(slug, category_id):
    xlsx = f"{slug}.xlsx"   # укажи свой путь

    # читаем A,B начиная с 10-й строки
    wb = load_workbook(xlsx, data_only=True, read_only=True)
    ws = wb.active  # если лист один; иначе wb["ИмяЛиста"]

    rows = []
    for r in ws.iter_rows(min_row=10, max_col=2, values_only=True):
        sku, price = r
        if sku is None or price is None:
            continue
        rows.append((str(sku).strip(), str(price), category_id, slug))

    with open_db() as conn:
        with conn.cursor() as cur:
            cur.executemany(
                """
                INSERT INTO tr_products_raw(sku, price, category_id, category_slug)
                VALUES (%s, %s, %s, %s)
                ON CONFLICT (sku) DO UPDATE SET price = EXCLUDED.price, category_id = EXCLUDED.category_id, category_slug = EXCLUDED.category_slug
                """,
                rows
            )
        conn.commit()

    print(f"Готово. Загружено строк: {len(rows)}")

def get_categories():
    with open_db() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT id, slug FROM tr_categories where is_leaf is true order by id asc")
            return cur.fetchall()

def etl():
    categories = get_categories()
    for i, category in enumerate(categories):
        print(f"Обрабатываем категорию {i+1}/{len(categories)}: {category[1]} ({category[0]})")
        try:
            get_xlsx(category[1])
            upsert_products_from_xlsx(category[1], category[0])
            print(f"✓ Успешно обработана: {category[1]}")
            # Задержка между запросами
            time.sleep(2)
        except Exception as e:
            print(f"✗ Ошибка при обработке {category[1]}: {e}")
            # При ошибке SSL попробуем подождать подольше
            if "SSL" in str(e) or "ssl" in str(e):
                print("SSL ошибка - ждем 30 секунд...")
                time.sleep(30)
            continue

etl()
# categories_raw = get_data()
# upsert_categories_from_json(categories_raw)
# slug = 'svarochnaja_aljuminievaja_provoloka'
# category_id = 1346
# get_xlsx(slug)
# upsert_products_from_xlsx(slug, category_id)